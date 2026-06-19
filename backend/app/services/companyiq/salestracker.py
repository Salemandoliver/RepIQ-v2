"""Sales Tracker — order-history source for CompanyIQ (replaces the old NetSuite feed).

Reads a locally-cached SharePoint Excel workbook (BTLB Sales Tracker.xlsx). Order history
is spread across monthly tabs named like "May 26 - Total LB MTD"; the order date comes
from the tab name. The workbook is parsed once and indexed in memory by normalised
company name, refreshed on a TTL. Any failure degrades to {"available": False} so the
panel never blocks.

Per-tab layout (0-indexed columns, header on row 3, data from row 4):
  0 Sales Agent (forward-filled — blank on rows after the first per agent)
  1 Company Name      2 Product Description    5 GM (£)
  7 Mobile  8 Cloud  9 Connectivity   11 Order Placed? ("y"/"n"/None)
Order value = cols 7+8+9; product category derived from which of those is non-zero.
"""
from __future__ import annotations

import base64
import io
import logging
import os
import re
import threading
import time
from urllib.parse import parse_qs, urlparse

import httpx

from ...config import settings
from .mastersheet import _norm_name

_BROWSER_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

log = logging.getLogger("calliq.companyiq.salestracker")

_lock = threading.Lock()
_st = {"loaded_at": 0.0, "by_name": {}, "tabs": 0, "rows": 0, "error": None}

_MONTHS = {m.lower(): i for i, m in enumerate(
    ["", "january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"])}
_ABBR = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _parse_tab_date(tab_name: str | None):
    """'May 26 - Total LB MTD' -> (2026, 5); None if unparseable."""
    m = re.match(r"\s*([A-Za-z]+)\s+(\d{2,4})", tab_name or "")
    if not m:
        return None
    month = _MONTHS.get(m.group(1).lower())
    if not month:
        return None
    yr = int(m.group(2))
    if yr < 100:
        yr += 2000
    return (yr, month)


def _fmt_period(period) -> str | None:
    if not period:
        return None
    yr, month = period
    return f"{_ABBR[month]} {yr}"


def _num(v) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)) or 0)
    except ValueError:
        return 0.0


def _tidy(n: float):
    return int(n) if float(n).is_integer() else round(n, 2)


def _cell(row, i):
    return row[i] if (i is not None and i < len(row)) else None


def _col_map(header_row) -> dict:
    """Map logical fields to column indices by HEADER NAME, so it works whether or not a
    tab has the newer 'Order Placed?' column (older tabs lack it)."""
    low = [str(h).strip().lower() if h is not None else "" for h in header_row]

    def find(*names):
        for i, h in enumerate(low):          # exact header match first
            if h in names:
                return i
        for i, h in enumerate(low):          # then substring
            if h and any(n in h for n in names):
                return i
        return None

    return {
        "agent": find("sales agent", "agent"),
        "company": find("company name", "company"),
        "product": find("product description", "product"),
        "gm": find("gm"),
        "mobile": find("mobile"),
        "cloud": find("cloud"),
        "connectivity": find("connectivity", "broadband"),
        "other": find("other"),
        "placed": find("order placed?", "order placed", "placed"),
    }


def _guestaccess_url(share_url: str) -> str | None:
    """Turn an 'anyone with the link' SharePoint share into a direct guestaccess download
    URL, e.g. .../personal/<user>/_layouts/15/guestaccess.aspx?e=<e>&download=1&share=<token>."""
    p = urlparse(share_url)
    # strip the share-redirect prefix, e.g. "/:x:/g/" -> "/"
    path = re.sub(r"^/:[a-zA-Z]:/[a-zA-Z]/", "/", p.path)
    parts = [x for x in path.split("/") if x]
    if len(parts) < 2:
        return None
    token = parts[-1]                       # the IQ... share token
    site = "/" + "/".join(parts[:-1])       # /personal/<user>  (or /sites/<site>)
    e = parse_qs(p.query).get("e", [""])[0]
    url = (f"https://{p.netloc}{site}/_layouts/15/guestaccess.aspx"
           f"?download=1&share={token}")
    if e:
        url += f"&e={e}"
    return url


def _download_xlsx(url: str) -> bytes:
    """GET an xlsx with a browser UA. Raises if the response is HTML (expired link)."""
    r = httpx.get(url, headers={"User-Agent": _BROWSER_UA}, timeout=60,
                  follow_redirects=True)
    r.raise_for_status()
    content = r.content
    if content[:2] != b"PK":   # .xlsx is a zip; an HTML login/expiry page would not be
        raise ValueError("share link did not return a file (expired or login page)")
    return content


def _graph_token() -> str:
    r = httpx.post(
        f"https://login.microsoftonline.com/{settings.ms_tenant_id}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": settings.ms_client_id,
              "client_secret": settings.ms_client_secret,
              "scope": "https://graph.microsoft.com/.default"}, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _fetch_sharepoint_bytes(share_url: str) -> bytes:
    """Download a workbook from a SharePoint/OneDrive sharing link via Microsoft Graph."""
    enc = "u!" + base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    r = httpx.get(f"https://graph.microsoft.com/v1.0/shares/{enc}/driveItem/content",
                  headers={"Authorization": f"Bearer {_graph_token()}"},
                  timeout=60, follow_redirects=True)
    r.raise_for_status()
    return r.content


class SalesTracker:
    @property
    def path(self) -> str:
        return settings.sales_tracker_xlsx

    @property
    def configured(self) -> bool:
        return bool(settings.sales_tracker_url or settings.sales_tracker_share_url) or \
            (bool(self.path) and os.path.exists(self.path))

    def _open_workbook(self):
        """Load the workbook: anonymous share download (preferred), then Graph, then file."""
        import openpyxl
        if settings.sales_tracker_url:
            dl = _guestaccess_url(settings.sales_tracker_url)
            if not dl:
                raise ValueError("could not parse SALES_TRACKER_URL share link")
            data = _download_xlsx(dl)
            return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if settings.sales_tracker_share_url:
            data = _fetch_sharepoint_bytes(settings.sales_tracker_share_url)
            return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return openpyxl.load_workbook(self.path, read_only=True, data_only=True)

    def _load(self, force: bool = False) -> None:
        with _lock:
            fresh = _st["loaded_at"] and (time.time() - _st["loaded_at"] < settings.sales_tracker_ttl)
            if fresh and not force:
                return
            if not self.configured:
                _st["error"] = "no Sales Tracker source configured"
                log.warning("SalesTracker not configured (set SALES_TRACKER_URL)")
                return
            try:
                wb = self._open_workbook()
            except Exception as e:
                _st["error"] = f"{type(e).__name__}: {str(e)[:300]}"
                log.warning("SalesTracker open failed: %s", e)
                return
            by_name, total = {}, 0
            matched, skipped, per_tab = [], [], []
            sample = None
            try:
                for ws in wb.worksheets:
                    title = ws.title or ""
                    # lenient: tolerates en-dash/hyphen and casing variations in tab names
                    if "total lb mtd" not in title.lower():
                        skipped.append(title)
                        continue
                    matched.append(title)
                    period = _parse_tab_date(title)
                    rows = list(ws.iter_rows(values_only=True))
                    # Find the header row (the one containing "Company Name"); its position
                    # and column order differ between older and newer tabs.
                    hidx = next((i for i, row in enumerate(rows[:8]) if row and any(
                        str(c).strip().lower() == "company name"
                        for c in row if c is not None)), None)
                    if hidx is None:
                        per_tab.append(f"{title}=0(no header)")
                        continue
                    cols = _col_map(rows[hidx])
                    has_placed = cols["placed"] is not None
                    agent = None
                    tab_n = 0
                    for row in rows[hidx + 1:]:
                        if row is None:
                            continue
                        a = _cell(row, cols["agent"])
                        if a not in (None, ""):
                            agent = str(a).strip()             # forward-fill the agent
                        company = _cell(row, cols["company"])
                        if company in (None, ""):
                            continue
                        if has_placed:                          # newer tabs: only Order Placed = y
                            placed = str(_cell(row, cols["placed"]) or "").strip().lower()
                            if not placed.startswith("y"):
                                continue
                        mobile = _num(_cell(row, cols["mobile"]))
                        cloud = _num(_cell(row, cols["cloud"]))
                        conn = _num(_cell(row, cols["connectivity"]))
                        other = _num(_cell(row, cols["other"]))
                        value = mobile + cloud + conn + other
                        if not has_placed and value <= 0:
                            continue                            # older tabs: skip summary/blank rows
                        nonzero = [label for label, v in
                                   (("Mobile", mobile), ("Cloud / Telephony", cloud),
                                    ("Broadband", conn), ("Other", other)) if v]
                        product = _cell(row, cols["product"])
                        category = ("Bundle" if len(nonzero) > 1 else nonzero[0] if nonzero
                                    else (str(product).strip() if product else "Order"))
                        by_name.setdefault(_norm_name(str(company)), []).append({
                            "period": period,
                            "date": _fmt_period(period),
                            "product": category,
                            "productDescription": str(product).strip() if product else None,
                            "value": _tidy(value),
                            "gm": _tidy(_num(_cell(row, cols["gm"]))),
                            "rep": agent,
                            "status": "Confirmed",
                        })
                        total += 1
                        tab_n += 1
                    per_tab.append(f"{title}={tab_n}")
                    if tab_n == 0 and sample is None and len(rows) > hidx + 1:
                        sample = {"tab": title, "cols": cols,
                                  "rows": [[("" if c is None else str(c)[:16])
                                            for c in (r or [])[:14]]
                                           for r in rows[hidx + 1:hidx + 5]]}
            finally:
                wb.close()
            _st.update({"loaded_at": time.time(), "by_name": by_name, "tabs": len(matched),
                        "rows": total, "error": None, "perTab": per_tab, "skipped": skipped,
                        "sample": sample})
            log.info("SalesTracker loaded %d confirmed orders across %d order tabs (of %d sheets). "
                     "Per-tab: %s. Non-order tabs skipped: %s",
                     total, len(matched), len(matched) + len(skipped),
                     " | ".join(per_tab), " | ".join(skipped[:25]))
            if sample:
                log.info("SalesTracker sample of empty order tab %r (rows 1-8, cols 0-12): %s",
                         sample["tab"], sample["rows"])

    def orders_for(self, name: str | None) -> dict:
        """Order-history block for a company. {'available': False} if the file can't be read."""
        if not self.configured:
            return {"available": False}
        self._load()
        if not _st["loaded_at"]:
            return {"available": False}
        rows = _st["by_name"].get(_norm_name(name), [])
        ordered = sorted(rows, key=lambda o: o.get("period") or (0, 0), reverse=True)
        orders = [{"date": o["date"], "product": o["product"], "value": o["value"],
                   "rep": o["rep"], "status": o["status"], "gm": o.get("gm")}
                  for o in ordered[:50]]
        return {"available": True, "totalOrders": len(rows),
                "lastOrderDate": orders[0]["date"] if orders else None, "orders": orders}


sales_tracker = SalesTracker()
