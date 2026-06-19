"""SalesIQ Lead Tracker + Activity Tracker readers.

Each tracker is reached one of two ways (whichever is configured, Graph preferred):
  (a) Microsoft Graph (Sites.Selected) by file path — LEAD_TRACKER_PATH / ACTIVITY_TRACKER_PATH
      — used when the tenant blocks anonymous sharing (see graph.py).
  (b) Anonymous SharePoint share — LEAD_TRACKER_URL / ACTIVITY_TRACKER_URL.
Read with openpyxl; columns matched by header name so they tolerate reordering. Layouts are
calibrated from the debug dumps (lead_debug / activity_debug) once access works — the same
approach used for the Sales Tracker.
"""
from __future__ import annotations

import io
import logging
import threading
import time
from datetime import date, datetime

import openpyxl

from ...config import settings
from ..companyiq.salestracker import _download_xlsx, _guestaccess_url, _num
from . import graph

log = logging.getLogger("calliq.salesiq.trackers")


def refresh() -> None:
    """Force a re-read of every configured tracker now (SalesIQ Refresh button)."""
    if leads_configured():
        _load_leads(force=True)
    if activity_configured():
        _load_activity(force=True)
    if holiday_configured():
        _load_holiday(force=True)


def _open_book(url: str, path: str):
    """Load a tracker workbook. Graph (Sites.Selected) by file `path` is preferred when
    configured; otherwise the anonymous SharePoint share `url`."""
    if path and graph.configured():
        data = graph.download(path)
    else:
        data = _download_xlsx(_guestaccess_url(url))
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _pick(header, *names):
    low = [str(h).strip().lower() if h is not None else "" for h in header]
    for i, h in enumerate(low):
        if h in names:
            return i
    for i, h in enumerate(low):
        if h and any(n in h for n in names):
            return i
    return None


def _cell(row, i):
    return row[i] if (i is not None and i < len(row)) else None


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d %b %Y", "%d %B %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def _find_header(rows, keywords, need=1):
    """First of the top rows containing >= `need` of the keywords."""
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        low = [str(c).strip().lower() for c in row if c is not None]
        hits = sum(1 for k in keywords if any(k in c for c in low))
        if hits >= need:
            return i
    return None


def _sheet_samples(wb, n_rows=15, n_cols=16):
    out = []
    for ws in wb.worksheets:
        rows = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= n_rows:
                break
            rows.append([("" if c is None else str(c)[:20]) for c in list(row)[:n_cols]])
        out.append({"sheet": ws.title, "rows": rows})
    return out


_MONTHS = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
           "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
           "december": 12, "sept": 9, "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6,
           "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
_YEAR_RE = __import__("re").compile(r"(20\d\d)")


def _sheet_year_month(title: str):
    """('June 2026', 'May 2026- BDM', 'Sept 2025 - Lead Generation') -> (year, month)."""
    t = title.strip().lower()
    ym = _YEAR_RE.search(t)
    if not ym:
        return None
    year = int(ym.group(1))
    for name in sorted(_MONTHS, key=len, reverse=True):  # longest first (september before sep)
        if name in t:
            return (year, _MONTHS[name])
    return None


def _flag(v) -> bool:
    """Truthy 'Yes'/'1'/'Won' style cell, treating No/blank/dash as false."""
    return str(v or "").strip().lower() not in ("", "no", "n", "0", "false", "-", "0.0")


# ============================================================ Lead Tracker
_lead_lock = threading.Lock()
_leads = {"loaded_at": 0.0, "rows": [], "error": None, "debug": None}


def leads_configured() -> bool:
    return bool(settings.lead_tracker_url or
                (settings.lead_tracker_path and graph.configured()))


def _load_leads(force: bool = False) -> None:
    with _lead_lock:
        fresh = _leads["loaded_at"] and (time.time() - _leads["loaded_at"] < settings.sales_tracker_ttl)
        if fresh and not force:
            return
        if not leads_configured():
            _leads["error"] = "not configured"
            return
        try:
            wb = _open_book(settings.lead_tracker_url, settings.lead_tracker_path)
        except Exception as e:
            _leads["error"] = f"{type(e).__name__}: {str(e)[:300]}"
            log.warning("Lead tracker open failed: %s", e)
            return
        out, dbg = [], None
        try:
            for ws in wb.worksheets:
                # Monthly detail sheets only, e.g. "June 2026 - Lead Generation".
                if "lead generation" not in ws.title.strip().lower():
                    continue
                rows = list(ws.iter_rows(values_only=True))
                hidx = _find_header(rows, ["company", "generator", "receiver", "date", "signed"], need=2)
                if hidx is None:
                    continue
                hdr = rows[hidx]
                cols = {
                    "bc": _pick(hdr, "generator", "business creator", "creator", "created by", "lead gen"),
                    "rep": _pick(hdr, "receiver", "received by", "passed to", "assigned", "closer"),
                    "company": _pick(hdr, "company name", "company", "customer", "account"),
                    "leadType": _pick(hdr, "lead type", "product"),
                    "date": _pick(hdr, "date received", "lead date", "date created", "date"),
                    "signed": _pick(hdr, "signed", "won"),
                    "f2f": _pick(hdr, "f2f", "face to face"),
                    "rejected": _pick(hdr, "rejected", "lost", "declined"),
                }
                if cols["company"] is None:
                    continue
                if dbg is None:
                    dbg = {"sheet": ws.title, "headerRow": hidx + 1, "cols": cols,
                           "sample": [[("" if c is None else str(c)[:20]) for c in (r or [])[:12]]
                                      for r in rows[hidx:hidx + 6]]}
                for row in rows[hidx + 1:]:
                    if not row:
                        continue
                    company = _cell(row, cols["company"])
                    if company in (None, ""):
                        continue
                    signed = _flag(_cell(row, cols["signed"]))
                    rejected = _flag(_cell(row, cols["rejected"]))
                    f2f = _flag(_cell(row, cols["f2f"]))
                    status = "Won" if signed else "Rejected" if rejected else "In Progress"
                    out.append({
                        "bc": str(_cell(row, cols["bc"]) or "").strip() or None,
                        "rep": str(_cell(row, cols["rep"]) or "").strip() or None,
                        "company": str(company).strip(),
                        "leadType": str(_cell(row, cols["leadType"]) or "").strip() or None,
                        "date": _to_date(_cell(row, cols["date"])),
                        "signed": signed, "f2f": f2f, "rejected": rejected,
                        "status": status, "outcome": None,
                    })
        finally:
            wb.close()
        _leads.update({"loaded_at": time.time(), "rows": out, "error": None, "debug": dbg})
        log.info("Lead tracker loaded %d leads", len(out))


def lead_rows():
    _load_leads()
    return _leads["rows"]


def lead_debug():
    if not leads_configured():
        return {"configured": False}
    _load_leads(force=True)
    try:
        wb = _open_book(settings.lead_tracker_url, settings.lead_tracker_path)
        samples = _sheet_samples(wb)
        wb.close()
    except Exception as e:
        samples = f"{type(e).__name__}: {str(e)[:200]}"
    return {"configured": True, "count": len(_leads["rows"]), "error": _leads["error"],
            "detected": _leads["debug"], "sheets": samples}


# ============================================================ Activity Tracker
_act_lock = threading.Lock()
_act = {"loaded_at": 0.0, "rows": [], "error": None, "debug": None}


def activity_configured() -> bool:
    return bool(settings.activity_tracker_url or
                (settings.activity_tracker_path and graph.configured()))


def _load_activity(force: bool = False) -> None:
    with _act_lock:
        fresh = _act["loaded_at"] and (time.time() - _act["loaded_at"] < settings.sales_tracker_ttl)
        if fresh and not force:
            return
        if not activity_configured():
            _act["error"] = "not configured"
            return
        try:
            wb = _open_book(settings.activity_tracker_url, settings.activity_tracker_path)
        except Exception as e:
            _act["error"] = f"{type(e).__name__}: {str(e)[:300]}"
            log.warning("Activity tracker open failed: %s", e)
            return
        out, dbg = [], None
        try:
            for ws in wb.worksheets:
                # One sheet per sales month: "June 2026" (Value/Volume) + "June 2026- BDM".
                ym = _sheet_year_month(ws.title)
                if not ym:
                    continue
                year, month = ym
                is_bdm = "bdm" in ws.title.strip().lower()
                rows = list(ws.iter_rows(values_only=True))
                hidx = _find_header(rows, ["sales agent", "opps", "new opps", "dials", "week"], need=2)
                if hidx is None:
                    continue
                hdr = rows[hidx]
                cols = {
                    "name": _pick(hdr, "sales agent", "rep name", "name", "agent"),
                    "team": _pick(hdr, "team"),
                    "date": _pick(hdr, "date"),
                    "atWork": _pick(hdr, "at work", "worked", "in?"),
                    "dials": _pick(hdr, "dials", "connected calls"),
                    "talk": _pick(hdr, "talk time", "talk", "minutes", "mins"),
                    "f2f": _pick(hdr, "f2f meeting", "f2f", "face to face"),
                    "opps": _pick(hdr, "opps created", "new opps", "opportunities", "opps", "opp"),
                    "leads": _pick(hdr, "leads sent", "leads"),
                }
                if cols["name"] is None:
                    continue
                if dbg is None:
                    dbg = {"sheet": ws.title, "headerRow": hidx + 1, "cols": cols,
                           "sample": [[("" if c is None else str(c)[:20]) for c in (r or [])[:14]]
                                      for r in rows[hidx:hidx + 6]]}
                for row in rows[hidx + 1:]:
                    if not row:
                        continue
                    nm = _cell(row, cols["name"])
                    if nm in (None, ""):
                        continue
                    agent = str(nm).strip()
                    if agent.lower().startswith("spare"):  # placeholder slots
                        continue
                    # Skip per-agent rollup rows (Week Total / Month Total / Monthly Avg) so
                    # daily figures aren't double/triple-counted; keep only real daily rows.
                    dval = str(_cell(row, cols["date"]) or "").strip().lower()
                    if not dval or any(k in dval for k in ("total", "avg", "average", "subtotal")):
                        continue
                    out.append({
                        "year": year, "month": month, "agent": agent, "isBdm": is_bdm,
                        "team": str(_cell(row, cols["team"]) or "").strip() or None,
                        "atWork": _flag(_cell(row, cols["atWork"])),
                        "dials": _num(_cell(row, cols["dials"])),
                        "talkMins": _num(_cell(row, cols["talk"])),
                        "f2f": _num(_cell(row, cols["f2f"])),
                        "opps": _num(_cell(row, cols["opps"])),
                        "leads": _num(_cell(row, cols["leads"])),
                    })
        finally:
            wb.close()
        _act.update({"loaded_at": time.time(), "rows": out, "error": None, "debug": dbg})
        log.info("Activity tracker loaded %d rows", len(out))


def activity_rows():
    _load_activity()
    return _act["rows"]


def activity_for(year: int, month: int) -> list[dict]:
    """Per-agent totals (opps/f2f/dials/talk/leads) for one sales month."""
    agg: dict[str, dict] = {}
    for r in activity_rows():
        if r["year"] != year or r["month"] != month:
            continue
        a = agg.setdefault(r["agent"].lower(),
                           {"agent": r["agent"], "team": r.get("team"), "isBdm": r.get("isBdm"),
                            "opps": 0.0, "f2f": 0.0, "dials": 0.0, "talkMins": 0.0, "leads": 0.0})
        for k in ("opps", "f2f", "dials", "talkMins", "leads"):
            a[k] += r.get(k) or 0
    return list(agg.values())


def activity_debug():
    if not activity_configured():
        return {"configured": False}
    _load_activity(force=True)
    try:
        wb = _open_book(settings.activity_tracker_url, settings.activity_tracker_path)
        samples = _sheet_samples(wb)
        wb.close()
    except Exception as e:
        samples = f"{type(e).__name__}: {str(e)[:200]}"
    return {"configured": True, "count": len(_act["rows"]), "error": _act["error"],
            "detected": _act["debug"], "sheets": samples}


# ============================================================ Holiday Tracker
_hol_lock = threading.Lock()
_hol = {"loaded_at": 0.0, "rows": [], "error": None, "debug": None}

# Cell codes -> (label, is_half). "B"/bank holiday is company-wide, not an absence (skipped).
HOLIDAY_CODES = {
    "h": ("Holiday", False), "h1": ("Half day (am)", True), "h2": ("Half day (pm)", True),
    "hd": ("Half day", True), "s": ("Sick", False), "s1": ("Sick (am)", True),
    "s2": ("Sick (pm)", True), "c": ("Compassionate", False), "n": ("Custom leave", False),
    "n2": ("Custom leave", False), "u": ("Unpaid leave", False),
}
_HOLIDAY_SKIP = {"b", "bh", "bank", ""}


def holiday_configured() -> bool:
    return bool(settings.holiday_tracker_url or
                (settings.holiday_tracker_path and graph.configured()))


def _as_day(v):
    try:
        d = int(float(v))
        return d if 1 <= d <= 31 else None
    except (TypeError, ValueError):
        return None


def _load_holiday(force: bool = False) -> None:
    with _hol_lock:
        fresh = _hol["loaded_at"] and (time.time() - _hol["loaded_at"] < settings.sales_tracker_ttl)
        if fresh and not force:
            return
        if not holiday_configured():
            _hol["error"] = "not configured"
            return
        try:
            wb = _open_book(settings.holiday_tracker_url, settings.holiday_tracker_path)
        except Exception as e:
            _hol["error"] = f"{type(e).__name__}: {str(e)[:300]}"
            log.warning("Holiday tracker open failed: %s", e)
            return
        out, dbg, cal = [], None, {}
        try:
            from datetime import date as _date
            for ws in wb.worksheets:
                ym = _sheet_year_month(ws.title)        # "June 2026" -> (2026, 6); "Totals" -> None
                if not ym:
                    continue
                year, month = ym
                rows = list(ws.iter_rows(values_only=True))
                # The day-number row: the first row whose cells (from col 2) are mostly 1..31.
                day_idx, col_day = None, {}
                for i, row in enumerate(rows[:8]):
                    if not row:
                        continue
                    days = {j: _as_day(c) for j, c in enumerate(row) if j >= 2 and _as_day(c)}
                    if len(days) >= 10:
                        day_idx, col_day = i, days
                        break
                if day_idx is None:
                    continue
                if dbg is None:
                    dbg = {"sheet": ws.title, "dayRow": day_idx + 1, "cols": len(col_day)}
                # Calendar day headers (sorted), with weekday + weekend flag.
                cal_days = []
                for j, day in sorted(col_day.items(), key=lambda kv: kv[1]):
                    try:
                        dt = _date(year, month, day)
                    except ValueError:
                        continue
                    cal_days.append({"day": day, "weekday": dt.strftime("%a"),
                                     "weekend": dt.weekday() >= 5})
                cal_people = []
                for row in rows[day_idx + 2:]:           # skip day row + weekday row
                    if not row:
                        continue
                    name = _cell(row, 0)
                    if name in (None, "") or not str(name).strip():
                        continue
                    name = str(name).strip()
                    cells = {}
                    for j, day in col_day.items():
                        raw = _cell(row, j)
                        code = str(raw or "").strip()
                        if not code:
                            continue
                        cells[day] = code.upper()        # keep all marks incl. B (bank holiday)
                        if code.lower() in _HOLIDAY_SKIP:
                            continue
                        label, half = HOLIDAY_CODES.get(code.lower(), (code, False))
                        try:
                            d = _date(year, month, day)
                        except ValueError:
                            continue
                        out.append({"name": name, "date": d, "code": code,
                                    "label": label, "half": half})
                    cal_people.append({"name": name.title(), "cells": cells})
                cal[(year, month)] = {"days": cal_days, "people": cal_people}
        finally:
            wb.close()
        _hol.update({"loaded_at": time.time(), "rows": out, "error": None,
                     "debug": dbg, "cal": cal})
        log.info("Holiday tracker loaded %d absence-days, %d month grids", len(out), len(cal))


def holiday_calendar(year: int, month: int) -> dict:
    """Full month grid (all employees × days with codes) for the calendar popup."""
    if not holiday_configured():
        return {"connected": False}
    _load_holiday()
    grid = (_hol.get("cal") or {}).get((year, month))
    if not grid:
        return {"connected": True, "found": False, "year": year, "month": month,
                "days": [], "people": []}
    return {"connected": True, "found": True, "year": year, "month": month,
            "days": grid["days"], "people": grid["people"]}


def holiday_rows():
    _load_holiday()
    return _hol["rows"]


def holiday_debug():
    if not holiday_configured():
        return {"configured": False}
    _load_holiday(force=True)
    try:
        wb = _open_book(settings.holiday_tracker_url, settings.holiday_tracker_path)
        samples = _sheet_samples(wb)
        wb.close()
    except Exception as e:
        samples = f"{type(e).__name__}: {str(e)[:200]}"
    return {"configured": True, "count": len(_hol["rows"]), "error": _hol["error"],
            "detected": _hol["debug"], "sheets": samples}
