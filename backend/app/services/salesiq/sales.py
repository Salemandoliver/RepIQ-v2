"""SalesIQ Sales Tracker reader.

Reuses the working Sales Tracker download (SALES_TRACKER_URL share link) and parses the
monthly "{Mon YY} - Total LB MTD" sheets into orders keyed by (year, month) — including
PENDING orders (Order Placed? = N) and sales-week tags, by header-name column detection so
it copes with the old/new tab layouts. Cached in memory on the same TTL as CompanyIQ.
"""
from __future__ import annotations

import io
import logging
import os
import threading
import time

import openpyxl

from ...config import settings
from ..companyiq.salestracker import (_cell, _download_xlsx, _fetch_sharepoint_bytes,
                                       _guestaccess_url, _num, _parse_tab_date)

log = logging.getLogger("calliq.salesiq.sales")

_lock = threading.Lock()
_state = {"loaded_at": 0.0, "by_period": {}, "months": [], "error": None}


def configured() -> bool:
    return bool(settings.sales_tracker_url or settings.sales_tracker_share_url
                or (settings.sales_tracker_xlsx and os.path.exists(settings.sales_tracker_xlsx)))


def _open_workbook():
    if settings.sales_tracker_url:
        data = _download_xlsx(_guestaccess_url(settings.sales_tracker_url))
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if settings.sales_tracker_share_url:
        data = _fetch_sharepoint_bytes(settings.sales_tracker_share_url)
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if settings.sales_tracker_xlsx and os.path.exists(settings.sales_tracker_xlsx):
        return openpyxl.load_workbook(settings.sales_tracker_xlsx, read_only=True, data_only=True)
    raise ValueError("Sales Tracker not configured (set SALES_TRACKER_URL)")


def _cols(header) -> dict:
    low = [str(h).strip().lower() if h is not None else "" for h in header]

    def find(*names):
        for i, h in enumerate(low):
            if h in names:
                return i
        for i, h in enumerate(low):
            if h and any(n in h for n in names):
                return i
        return None

    return {
        "agent": find("sales agent", "agent"),
        "company": find("company name", "company"),
        "product": find("product description", "product"),
        "split_with": find("split with"),
        "split_pct": find("split %", "split%"),
        "gm": find("gm"),
        "cobra": find("cobra gm", "cobra"),
        "mobile": find("mobile"),
        "cloud": find("cloud"),
        "connectivity": find("connectivity", "broadband"),
        "other": find("other"),
        "placed": find("order placed?", "order placed", "placed"),
    }


def _txt(v):
    return str(v).strip() if v not in (None, "") else None


def _load(force: bool = False) -> None:
    with _lock:
        fresh = _state["loaded_at"] and (time.time() - _state["loaded_at"] < settings.sales_tracker_ttl)
        if fresh and not force:
            return
        if not configured():
            _state["error"] = "Sales Tracker not configured"
            return
        try:
            wb = _open_workbook()
        except Exception as e:
            _state["error"] = f"{type(e).__name__}: {str(e)[:300]}"
            log.warning("SalesIQ sales open failed: %s", e)
            return
        by_period, months = {}, []
        try:
            for ws in wb.worksheets:
                title = ws.title or ""
                if "total lb mtd" not in title.lower():
                    continue
                period = _parse_tab_date(title)
                if not period:
                    continue
                rows = list(ws.iter_rows(values_only=True))
                hidx = next((i for i, row in enumerate(rows[:8]) if row and any(
                    str(c).strip().lower() == "company name"
                    for c in row if c is not None)), None)
                if hidx is None:
                    continue
                cols = _cols(rows[hidx])
                has_placed = cols["placed"] is not None
                # Orders can span several rows: the agent + company appear only on the first
                # line of an order, continuation lines (extra products) leave both blank. So we
                # forward-fill agent AND company, and keep every line that carries a product or
                # value. Each order is assigned to the sales week of the subtotal row that
                # *follows* it, so we buffer lines and flush them when a "Week N" row appears.
                orders, buf = [], []
                agent, company, last_wk = None, None, 0

                def _wk_num(label: str) -> int:
                    d = "".join( c for c in label if c.isdigit())
                    return int(d) if d else 0

                def _flush(label: str):
                    for o in buf:
                        o["week"] = label
                    orders.extend(buf)
                    buf.clear()

                for row in rows[hidx + 1:]:
                    if row is None:
                        continue
                    a = _cell(row, cols["agent"])
                    if a not in (None, ""):
                        new_agent = str(a).strip()
                        if new_agent != agent:
                            if buf:                              # close out prior agent
                                _flush(f"Week {last_wk + 1}" if last_wk else "Week 1")
                            agent, company, last_wk = new_agent, None, 0
                    sp = _cell(row, cols["split_pct"])
                    if sp and str(sp).strip().lower().startswith("week"):
                        lbl = str(sp).strip()                    # week subtotal row
                        last_wk = _wk_num(lbl) or last_wk
                        _flush(lbl)
                        company = None                           # don't let an order span weeks
                        continue
                    c = _cell(row, cols["company"])
                    if c not in (None, ""):
                        company = str(c).strip()
                    product = _txt(_cell(row, cols["product"]))
                    mob = _num(_cell(row, cols["mobile"]))
                    cl = _num(_cell(row, cols["cloud"]))
                    cn = _num(_cell(row, cols["connectivity"]))
                    ot = _num(_cell(row, cols["other"]))
                    sov = mob + cl + cn + ot
                    gm = _num(_cell(row, cols["gm"]))
                    if not product and sov <= 0 and gm <= 0:
                        continue                                 # blank / separator / total row
                    if not company:
                        continue                                 # can't attribute without a company
                    placed = (str(_cell(row, cols["placed"]) or "").strip().lower().startswith("y")
                              if has_placed else True)
                    buf.append({
                        "agent": agent,
                        "company": company,
                        "product": product,
                        "splitWith": _txt(_cell(row, cols["split_with"])),
                        "splitPct": _cell(row, cols["split_pct"]),
                        "gm": round(gm, 2),
                        "cobraGm": round(_num(_cell(row, cols["cobra"])), 2),
                        "mobile": round(mob, 2), "cloud": round(cl, 2),
                        "connectivity": round(cn, 2), "other": round(ot, 2),
                        "sov": round(sov, 2),
                        "placed": placed, "week": f"Week {last_wk + 1}" if last_wk else "Week 1",
                    })
                if buf:                                          # trailing orders (no following subtotal)
                    _flush(f"Week {last_wk + 1}" if last_wk else "Week 1")
                by_period.setdefault(period, []).extend(orders)
                months.append({"period": list(period), "tab": title, "orders": len(orders)})
        finally:
            wb.close()
        _state.update({"loaded_at": time.time(), "by_period": by_period,
                       "months": months, "error": None})
        log.info("SalesIQ sales loaded %d monthly periods (%s)", len(by_period),
                 ", ".join(f"{m}/{y}={len(o)}" for (y, m), o in sorted(by_period.items())))


def refresh() -> None:
    """Force a re-read of the Sales Tracker now (used by the SalesIQ Refresh button)."""
    _load(force=True)


def orders_for(year: int, month: int) -> list[dict]:
    _load()
    return _state["by_period"].get((year, month), [])


def status() -> dict:
    _load()
    return {"configured": configured(), "months": _state["months"], "error": _state["error"]}


def master_dashboard_dump() -> dict:
    """First rows/cols of the 'Master Dashboard' sheet, to calibrate the manager figures."""
    try:
        wb = _open_workbook()
    except Exception as e:
        return {"error": f"{type(e).__name__}: {str(e)[:300]}"}
    try:
        title = next((ws.title for ws in wb.worksheets
                      if "master dashboard" in (ws.title or "").lower()), None)
        if not title:
            return {"error": "no 'Master Dashboard' sheet found",
                    "sheets": [w.title for w in wb.worksheets]}
        ws = wb[title]
        rows = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= 35:
                break
            rows.append([("" if c is None else str(c)[:22]) for c in list(row)[:16]])
        return {"sheet": title, "rows": rows}
    finally:
        wb.close()

