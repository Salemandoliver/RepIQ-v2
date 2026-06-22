"""NetSuite ERP Dump importer (brief §14.8, §14.11).

Brings existing orders into RepIQ from the NetSuite 'ERP Dump' export (one row per order *line*,
grouped by Document Number / SO#), so the platform runs in parallel with the live Excel/NetSuite
trackers until we're confident. Two-step like the HR import: **dry-run** (preview + unmatched
headers, nothing written) then **commit**.

Only the current financial year is imported — rows dated before the FY start (Mon 30 Mar 2026 for
FY27) are skipped, per the BT calendar. Idempotent by SO#: orders already in RepIQ are left untouched
so manual edits are never clobbered.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime

from sqlalchemy.orm import Session

from ...models import User
from ...services.salesiq.fincal import financial_year_start
from ...services.salesiq.roles import agent_matches
from .models import ORDER_STATUS, Customer, Order, OrderAgent, OrderLine
from .services import recompute_totals, stamp_financial_month, stamp_week

# Canonical keys → list of accepted header spellings (lower/stripped). Tolerant to NetSuite variants.
HEADER_SYNONYMS = {
    "so": ["document number", "document number (so#)", "document number (no.)", "so#", "so #",
           "so number", "so no", "order number", "order #", "order no", "sales order", "sales order #",
           "doc number", "document no", "transaction number", "tran id"],
    "date": ["date", "order date"],
    "company": ["company name", "company"],
    "le": ["le", "le code", "customer le"],
    "status": ["order status"],
    "status_changed": ["order status last changed date", "order status last changed"],
    "opp_id": ["opp id", "opportunity id"],
    "cancellation_date": ["cancellation date"],
    "main_order_number": ["main order number", "job number"],
    "sales_rep": ["sales rep", "primary sales rep", "rep"],
    "primary": ["primary", "primary (y/n)", "primary flag"],
    "sales_role": ["sales role name", "sales role", "sales role: name"],
    "admin_agent": ["admin agent"],
    "item": ["item name (grouped)", "item: name (grouped)", "item name", "item: name", "item"],
    "item_id": ["item id"],
    "group1": ["product group 1", "product group1"],
    "group2": ["product group 2", "product group2"],
    "klass": ["class (item) name", "class (item): name", "class", "class (item)"],
    "schedule5": ["item schedule 5", "item: schedule 5", "schedule 5", "schedule 5 area"],
    "contract_value": ["contract value"],
    "quantity": ["quantity", "qty"],
    "gm": ["total sov", "sov", "total bookings amount", "gm", "gross margin", "contribution amount (net)"],
    "contribution_amount": ["contribution amount (net)", "contribution amount"],
    "contribution_pct": ["contribution %", "contribution percent"],
    "agent1_split": ["agent 1 split", "agent1 split"],
    "agent2_split": ["agent 2 split", "agent2 split"],
    "agent3_split": ["agent 3 split", "agent3 split"],
    "dirty": ["sales team issue / dirty order?", "dirty order", "sales team issue"],
    "bt_commission_paid": ["bt commission paid", "item on finance"],
    "schedule5_check": ["schedule 5 check"],
}
_BOM = "﻿"


def _norm(h: str) -> str:
    return (h or "").replace(_BOM, "").strip().lower()


def _build_index(headers: list[str]) -> tuple[dict, list[str]]:
    """Map canonical key → column index. Returns (index, unmatched_headers)."""
    norm = [_norm(h) for h in headers]
    idx, used = {}, set()
    for key, spellings in HEADER_SYNONYMS.items():
        for sp in spellings:
            if sp in norm:
                idx[key] = norm.index(sp)
                used.add(norm.index(sp))
                break
    unmatched = [headers[i] for i in range(len(headers)) if i not in used and norm[i]]
    return idx, unmatched


def _parse_spreadsheetml(data: bytes) -> list[list]:
    """Parse the Microsoft 'XML Spreadsheet 2003' format (NetSuite's .xls export — actually XML)."""
    import xml.etree.ElementTree as ET
    NS = "{urn:schemas-microsoft-com:office:spreadsheet}"
    try:
        root = ET.fromstring(data.decode("utf-8-sig", errors="replace"))
    except ET.ParseError as e:
        raise ValueError(f"Could not parse the NetSuite Excel-XML file: {e}")
    table = next((t for t in root.iter(NS + "Table")), None)
    if table is None:
        return []
    out = []
    for row in table.findall(NS + "Row"):
        cells, col = [], 0
        for cell in row.findall(NS + "Cell"):
            idx_attr = cell.get(NS + "Index")          # sparse rows skip empty cells
            if idx_attr:
                target = int(idx_attr) - 1
                while col < target:
                    cells.append("")
                    col += 1
            d = cell.find(NS + "Data")
            cells.append(d.text if (d is not None and d.text is not None) else "")
            col += 1
        out.append(cells)
    return out


def _parse_html_table(data: bytes) -> list[list]:
    """Parse a simple HTML <table> export (NetSuite sometimes emits .xls as HTML)."""
    from html.parser import HTMLParser

    class _P(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows, self.cur, self.cell = [], None, None

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self.cur = []
            elif tag in ("td", "th"):
                self.cell = ""

        def handle_endtag(self, tag):
            if tag == "tr" and self.cur is not None:
                self.rows.append(self.cur)
                self.cur = None
            elif tag in ("td", "th") and self.cur is not None:
                self.cur.append((self.cell or "").strip())
                self.cell = None

        def handle_data(self, d):
            if self.cell is not None:
                self.cell += d

    p = _P()
    p.feed(data.decode("utf-8-sig", errors="replace"))
    return [r for r in p.rows if any(str(c).strip() for c in r)]


def all_rows(data: bytes, filename: str) -> list[list]:
    """Every row of the file — handles CSV, real XLSX, NetSuite Excel-XML (.xls) and HTML-table .xls."""
    name = (filename or "").lower()
    sniff = data[:4096].decode("utf-8", errors="replace").lower()
    if data[:2] == b"PK" or name.endswith((".xlsx", ".xlsm")):   # real zip-based xlsx
        try:
            import openpyxl
        except ImportError:
            raise ValueError("XLSX not supported on the server — please export the ERP Dump as CSV.")
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[c if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
    if "urn:schemas-microsoft-com:office:spreadsheet" in sniff:   # NetSuite Excel-XML .xls
        return _parse_spreadsheetml(data)
    if "<table" in sniff or "<tr" in sniff:                       # HTML-table .xls
        return _parse_html_table(data)
    text = data.decode("utf-8-sig", errors="replace")            # plain CSV
    return list(csv.reader(io.StringIO(text)))


def _all_spellings() -> set:
    s: set = set()
    for sp in HEADER_SYNONYMS.values():
        s.update(sp)
    return s


def _best_header_row(rows: list[list]) -> int:
    """NetSuite/Excel exports sometimes have a title or blank row above the real headers. Pick the
    row (within the first 12) that matches the most known column names."""
    spell = _all_spellings()
    best_i, best = 0, -1
    for i in range(min(12, len(rows))):
        sc = sum(1 for c in rows[i] if _norm(str(c)) in spell)
        if sc > best:
            best, best_i = sc, i
    return best_i


def parse_file(data: bytes, filename: str) -> tuple[list[str], list[list]]:
    """Return (headers, rows) — auto-detecting the header row past any title/blank rows."""
    rows = all_rows(data, filename)
    if not rows:
        return [], []
    hi = _best_header_row(rows)
    return [str(h) for h in rows[hi]], rows[hi + 1:]


def _cell(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    v = row[i]
    return v if (v is not None and str(v).strip() != "") else None


def _to_date(v) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)) or 0)
    except ValueError:
        return 0.0


def _yn(v) -> bool:
    return _norm(str(v or "")) in ("y", "yes", "true", "1", "paid", "✓", "x")


def _s5check(v):
    t = _norm(str(v or ""))
    if not t:
        return None
    if "not on" in t or "not_on" in t:
        return "not_on"
    if "incorrect" in t:
        return "on_incorrect"
    if "correct" in t:
        return "on_correct"
    return None


def _status_code(text) -> str:
    t = _norm(str(text or ""))
    t = re.sub(r"^[a-z]\.\s*", "", t)        # strip "I. " prefix
    if not t:
        return "O"
    for code, label in ORDER_STATUS.items():
        if t == label.lower() or label.lower() in t or t in label.lower():
            return code
    if "issue" in t and "pay" in t:
        return "P"
    if "non" in t and "commission" in t:
        return "N"
    if "cancel" in t:
        return "M"
    return "O"


def _group(data: bytes, filename: str, floor: date):
    """Parse → group rows by SO# → per-order dict. Returns (orders, unmatched_headers, skipped_old)."""
    headers, rows = parse_file(data, filename)
    idx, unmatched = _build_index(headers)
    if "so" not in idx:
        found = ", ".join(str(h).strip() for h in headers if str(h).strip())[:600]
        raise ValueError("Could not find a Document Number / SO# column. Columns found in the file: "
                         + (found or "(none — is the first sheet the data sheet?)")
                         + ". Tell Claude which column is the SO/order number and it'll be mapped.")
    orders: dict[str, dict] = {}
    skipped_old = 0
    for row in rows:
        so = _cell(row, idx, "so")
        if not so:
            continue
        so = str(so).strip()
        d = _to_date(_cell(row, idx, "date"))
        if d and d < floor:
            skipped_old += 1
            continue
        o = orders.setdefault(so, {
            "so": so, "date": d, "company": _cell(row, idx, "company"),
            "le": _cell(row, idx, "le"), "opp_id": _cell(row, idx, "opp_id"),
            "status": _status_code(_cell(row, idx, "status")),
            "cancellation_date": _to_date(_cell(row, idx, "cancellation_date")),
            "main_order_number": _cell(row, idx, "main_order_number"),
            "admin_agent": _cell(row, idx, "admin_agent"),
            "lines": [], "agents": {},
        })
        o["lines"].append({
            "item": _cell(row, idx, "item") or "", "group1": _cell(row, idx, "group1"),
            "group2": _cell(row, idx, "group2"), "klass": _cell(row, idx, "klass"),
            "schedule5": _cell(row, idx, "schedule5"),
            "contract_value": _to_float(_cell(row, idx, "contract_value")),
            "quantity": int(_to_float(_cell(row, idx, "quantity")) or 1),
            "gm": _to_float(_cell(row, idx, "gm")),
            "bt_commission_paid": _yn(_cell(row, idx, "bt_commission_paid")),
            "schedule5_check": _s5check(_cell(row, idx, "schedule5_check")),
        })
        rep = _cell(row, idx, "sales_rep")
        if rep:
            a = o["agents"].setdefault(str(rep), {
                "name": str(rep), "sales_role": _cell(row, idx, "sales_role"),
                "is_primary": _norm(str(_cell(row, idx, "primary") or "")) in ("y", "yes", "true", "1"),
                "contribution_pct": _to_float(_cell(row, idx, "contribution_pct")),
            })
    return list(orders.values()), unmatched, skipped_old


def analyze(db: Session, data: bytes, filename: str, floor: date | None = None) -> dict:
    floor = floor or financial_year_start()
    orders, unmatched, skipped_old = _group(data, filename, floor)
    existing = {n for (n,) in db.query(Order.order_number).all()}
    preview, new, dupes = [], 0, 0
    for o in orders:
        is_new = o["so"] not in existing
        new += 1 if is_new else 0
        dupes += 0 if is_new else 1
        preview.append({"so": o["so"], "date": o["date"].isoformat() if o["date"] else None,
                        "company": o["company"], "status": o["status"], "lines": len(o["lines"]),
                        "agents": len(o["agents"]), "isNew": is_new})
    return {"floor": floor.isoformat(), "totalOrders": len(orders), "new": new, "duplicates": dupes,
            "skippedBeforeFY": skipped_old, "unmatchedHeaders": unmatched,
            "preview": preview[:200]}


_BATCH_SIZE = 50          # commit incrementally so a slow/timed-out import still persists + resumes


def commit(db: Session, data: bytes, filename: str, user: User, floor: date | None = None) -> dict:
    """Import orders. Idempotent by SO# (re-running skips orders already in RepIQ), so a retry after
    a timeout safely resumes. Commits in batches to keep transactions short and avoid HTTP timeouts."""
    floor = floor or financial_year_start()
    orders, _, skipped_old = _group(data, filename, floor)
    existing = {n for (n,) in db.query(Order.order_number).all()}
    batch = f"erp-{datetime.utcnow():%Y%m%d%H%M%S}"

    # Precompute lookups ONCE (the per-order queries were the slow part / cause of the hang).
    user_index = [(u.id, u.name, u.short_name) for u in db.query(User).all()]
    cust_by_le: dict = {}
    cust_by_name: dict = {}
    for c in db.query(Customer).filter(Customer.deleted_at.is_(None)).all():
        if c.le_code:
            cust_by_le[c.le_code.strip()] = c
        cust_by_name[(c.company_name or "").strip().lower()] = c

    def get_customer(le, name):
        le = (le or "").strip() or None
        name = (name or "").strip()
        if not le and not name:
            return None
        c = (cust_by_le.get(le) if le else None) or (cust_by_name.get(name.lower()) if name else None)
        if c:
            return c
        c = Customer(le_code=le, company_name=name or le)
        db.add(c)
        db.flush()
        if le:
            cust_by_le[le] = c
        cust_by_name[(c.company_name or "").lower()] = c
        return c

    agent_cache: dict = {}

    def match_user(agent_name):
        if agent_name in agent_cache:
            return agent_cache[agent_name]
        uid = None
        for u_id, nm, sn in user_index:
            if agent_matches(nm, agent_name) or (sn and agent_matches(sn, agent_name)):
                uid = u_id
                break
        agent_cache[agent_name] = uid
        return uid

    created, pending = 0, 0
    for od in orders:
        if od["so"] in existing:
            continue
        cust = get_customer(od["le"], od["company"] or "")
        o = Order(order_number=od["so"], order_date=od["date"] or date.today(),
                  status=od["status"], company_name=cust.company_name if cust else (od["company"] or ""),
                  le_code=cust.le_code if cust else od["le"], customer_id=cust.id if cust else None,
                  opp_id=od["opp_id"], main_order_number=od["main_order_number"],
                  cancellation_date=od["cancellation_date"],
                  order_cancelled=(od["status"] == "M"), source="import", import_batch=batch,
                  created_by_id=user.id)
        stamp_financial_month(o)
        stamp_week(o)
        db.add(o)
        db.flush()
        for i, ln in enumerate(od["lines"], start=1):
            db.add(OrderLine(order_id=o.id, line_no=i, item_name=ln["item"],
                             product_group1=ln["group1"], product_group2=ln["group2"],
                             schedule5_area=ln["schedule5"], contract_value=ln["contract_value"],
                             quantity=ln["quantity"], gm=ln["gm"],
                             bt_commission_paid=ln["bt_commission_paid"],
                             schedule5_check=ln["schedule5_check"]))
        for a in od["agents"].values():
            db.add(OrderAgent(order_id=o.id, user_id=match_user(a["name"]),
                              agent_name=a["name"], sales_role=a["sales_role"],
                              is_primary=a["is_primary"], contribution_pct=a["contribution_pct"]))
        db.flush()
        recompute_totals(o)
        created += 1
        pending += 1
        if pending >= _BATCH_SIZE:
            db.commit()
            pending = 0
    db.commit()
    return {"created": created, "skippedExisting": len(orders) - created,
            "skippedBeforeFY": skipped_old, "batch": batch}
