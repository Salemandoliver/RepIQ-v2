"""BT rate-card importer — loads the yearly 'BTLB Rate Card' (.xlsx) into the product catalogue.

Each row is a BT product with its Product Group 1/2, Product Area, Schedule 5 area, Cobra report and
the commission rate per reporting period (P1–P12). We upsert these into OrderProduct (keyed by the
SalesForce reference), stamp each with its BT targeting category (Data/Cloud/Mobile) and store the
current-period rate + ref in the product's metadata for the commission engine to use later.
"""
from __future__ import annotations

import io

from sqlalchemy.orm import Session

from .categories import bt_category
from .models import OrderProduct


def _norm(x) -> str:
    return str(x if x is not None else "").strip().lower()


def _v(row, i):
    if i is None or i >= len(row):
        return None
    v = row[i]
    return str(v).strip() if (v is not None and str(v).strip()) else None


def _rate(row, i):
    if i is None or i >= len(row):
        return None
    v = row[i]
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None          # e.g. "N/A - Bespoke"


def import_rate_card(db: Session, data: bytes, filename: str = "") -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if "rate card" in s.lower()), wb[wb.sheetnames[0]])
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The rate card sheet is empty.")

    hi = None
    for i, r in enumerate(rows[:20]):
        if any(_norm(c) == "product description" for c in r):
            hi = i
            break
    if hi is None:
        raise ValueError("Couldn't find the rate-card header row (need a 'Product Description' column).")
    header = [_norm(c) for c in rows[hi]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        for j, h in enumerate(header):
            if h and any(h.startswith(n) for n in names):
                return j
        return None

    c_ref = col("salesforce ref", "salesforce reference")
    c_desc = col("product description")
    c_g1 = col("product group 1")
    c_g2 = col("product group 2")
    c_area = col("product area")
    c_s5 = col("schedule 5 / non sched", "schedule 5", "schedule 5 / non schedule 5")
    c_cobra = col("cobra report", "cobra")
    c_rate = col("p1 2627", "p1")

    by_ref, by_name = {}, {}
    for p in db.query(OrderProduct).all():
        ref = (p.extra or {}).get("salesforceRef")
        if ref:
            by_ref[ref] = p
        by_name[_norm(p.name)] = p

    created = updated = 0
    seen_refs = set()
    for r in rows[hi + 1:]:
        desc = _v(r, c_desc)
        if not desc:
            continue
        ref = _v(r, c_ref)
        g1, g2, area = _v(r, c_g1), _v(r, c_g2), _v(r, c_area)
        s5, cobra = _v(r, c_s5), _v(r, c_cobra)
        rate = _rate(r, c_rate)
        cat = bt_category(g1, g2, area, s5)
        meta = {"salesforceRef": ref, "rate": rate, "btCategory": cat}
        if ref:
            seen_refs.add(ref)
        p = (by_ref.get(ref) if ref else None) or by_name.get(_norm(desc))
        if p:
            p.name, p.product_group1, p.product_group2 = desc, g1, g2
            p.product_class, p.schedule5_area, p.cobra, p.active = area, s5, cobra, True
            p.extra = {**(p.extra or {}), **meta}
            updated += 1
        else:
            db.add(OrderProduct(name=desc, product_group1=g1, product_group2=g2, product_class=area,
                                schedule5_area=s5, cobra=cobra, active=True, extra=meta))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "products": created + updated}
